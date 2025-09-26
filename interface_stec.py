import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from traitement_pipeline import lancer_pipeline_agencement
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as RLImage, PageBreak
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from os import makedirs, path
import tkinter.messagebox



chemin_fichier_selectionne = None  

def generer_template_excel():
    from openpyxl import Workbook
    wb = Workbook()

    
    feuilles = {
        "Extraction": ["URL extraction"],
        "Compétences": ["Compétences STEC"],
        "Prat": ["URL prat"],
        "NPP": ["URL NPP"],
        "Gares_criteres": ["URL Gares et critères STEC"]
    }

    for nom, colonnes in feuilles.items():
        ws = wb.create_sheet(title=nom)
        ws.append(colonnes)

   
    ws_aide = wb.create_sheet(title="Aide")
    aide_text = [
        ["Veillez fournir le ou les urls des fichier excels suivant dans les feuilles concernées"],
        ["faire attention à bien fournir des liens avec des \\\\ et pas \\ , exemple d'un lien url valide : c:\\\\Users\\\\Input\\\\donnees_prat.xlsx"],
        ["Important : si la syntaxe des colonnes n'est pas scrupulesement respectée, le programme ne marchera pas"],
        [""],
        ["Les colonnes indispensables ci-dessous doivent impérativement figurer dans leurs excels respectifs,"],
        ["d'autres colonnes peuvent y figurer aussi; elles seront simplement pas prises en compte"],
        [""],
        ["Fichier excel", "Les colonnes indispensables"],
        ["Extraction", '"Compétence", "Engin", "Caisse", "Kit", "Code tâche","date_debut", "emplacement wms entrée"'],
        ["NPP", '"Code article", "Code contenant"'],
        ["Prat", '"Nom du Praticable", "Surface", "Gerbage"'],
        ["Gares_criteres", '"gare", "duree min pour STEC", "duree min pour STEC si mm gare"'],
        ["Compétences", '"Compétences STEC"']
    ]

    for row in aide_text:
        ws_aide.append(row)

    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save("Template_STEC.xlsx")
    messagebox.showinfo("Succès", "Le template Excel a été généré : Template_STEC.xlsx")


def charger_fichier_excel():
    global chemin_fichier_selectionne
    fichier = filedialog.askopenfilename(
        title="Sélectionnez un fichier Excel",
        filetypes=[("Fichiers Excel", "*.xlsx")]
    )
    if fichier:
        chemin_fichier_selectionne = fichier
        messagebox.showinfo("Fichier chargé", f"Fichier sélectionné : {fichier}")

def exporter_pdf(positions, df_max_supports):
    
    if positions is None or df_max_supports is None:
        tkinter.messagebox.showwarning("Pas de données", "Veuillez d'abord lancer l'analyse STEC.")
        return

    pdf_path = "OUTPUT/rapport_agencement.pdf"
    img_path = "OUTPUT/agencement.png"

    if not path.exists(img_path):
        tkinter.messagebox.showerror("Erreur", "L'image d'agencement est introuvable.")
        return

    try:
        makedirs("OUTPUT", exist_ok=True)

        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        elements = []

        # Image
        img = RLImage(img_path, width=16 * cm, height=12 * cm)
        elements.append(img)
        elements.append(PageBreak())

        # Tableau
        tableau_data = [["Contenant", "Supports", "Surface (m²)"]]
        total_supports = 0
        total_surface = 0.0

        for _, row in df_max_supports.iterrows():
            support = row["nb supports après gerbage"]
            surface = round(row["surface après gerbage"], 2)
            tableau_data.append([row["Type support"], support, surface])
            total_supports += support
            total_surface += surface

        tableau_data.append(["TOTAL", total_supports, round(total_surface, 2)])

        table = Table(tableau_data, colWidths=[8*cm, 4*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (-1, -1), (-1, -1), colors.lightgrey)
        ]))

        elements.append(table)

        doc.build(elements)
        

    except Exception as e:
        tkinter.messagebox.showerror("Erreur PDF", f"Erreur lors de la génération du PDF : {e}")

def lancer_analyses():
    global chemin_fichier_selectionne
    if not chemin_fichier_selectionne:
        messagebox.showerror("Erreur", "Veuillez d'abord charger un fichier Excel")
        return
    try:
        positions, df_max_supports = lancer_pipeline_agencement(chemin_fichier_selectionne)
        messagebox.showinfo("Succès", "Analyses terminées. Résultats disponibles dans OUTPUT/")
        exporter_pdf(positions, df_max_supports)
    except Exception as e:
        messagebox.showerror("Erreur", str(e))

def lancer_interface():
    """Fonction principale pour lancer l'interface STEC"""
    root = ctk.CTk()
    root.title("STEC Tool")
    root.geometry("500x300")

    bouton_template = ctk.CTkButton(root, text="Générer Template Excel", command=generer_template_excel)
    bouton_template.pack(pady=10)

    bouton_charger = ctk.CTkButton(root, text="Charger un fichier Excel", command=charger_fichier_excel)
    bouton_charger.pack(pady=10)

    bouton_analyse = ctk.CTkButton(root, text="Lancer les analyses", command=lancer_analyses)
    bouton_analyse.pack(pady=20)

    root.mainloop()


if __name__ == "__main__":
    lancer_interface()


